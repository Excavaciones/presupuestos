# -*- coding: utf-8 -*-
"""
explorador_bc3.py  v2
Explorador de archivos BC3 con exportación a Excel y portapapeles.

Uso:  python explorador_bc3.py
      python explorador_bc3.py mi_archivo.bc3

Requiere: Python 3, tkinter (incluido en Windows), openpyxl
Instalar openpyxl si no lo tienes:  pip install openpyxl
"""

import re
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    TIENE_OPENPYXL = True
except ImportError:
    TIENE_OPENPYXL = False

from presupuesto import presupuesto


COLOR_NARANJA  = '#E87722'
COLOR_FONDO    = '#F5F5F5'
COLOR_BLANCO   = '#FFFFFF'
COLOR_GRIS     = '#DDDDDD'
COLOR_AZUL     = '#1A5276'
COLOR_VERDE    = '#1E8449'
COLOR_TEXTO    = '#222222'


class ExploradorBC3:
    def __init__(self, root):
        self.root = root
        self.root.title('Explorador BC3 — Excavaciones Arturo S.L.')
        self.root.configure(bg=COLOR_FONDO)
        self.root.geometry('1280x760')
        self.pres = None
        self.carrito = []
        self._orden_col = None      # columna actual de ordenación
        self._orden_asc = True      # True=ascendente, False=descendente
        self._filas_cache = []      # cache de filas sin ordenar
        self._construir_ui()
        self._estilo_tabla()

        if len(sys.argv) > 1:
            self._cargar_archivo(sys.argv[1])

    def _estilo_tabla(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Treeview', font=('Arial', 9), rowheight=22,
                        background=COLOR_BLANCO, fieldbackground=COLOR_BLANCO)
        style.configure('Treeview.Heading', font=('Arial', 9, 'bold'),
                        background=COLOR_NARANJA, foreground=COLOR_BLANCO)
        style.map('Treeview', background=[('selected', COLOR_AZUL)],
                  foreground=[('selected', COLOR_BLANCO)])

    def _construir_ui(self):
        # Barra superior
        barra = tk.Frame(self.root, bg=COLOR_NARANJA, pady=8)
        barra.pack(fill='x')
        tk.Label(barra, text='Explorador BC3', bg=COLOR_NARANJA,
                 fg=COLOR_BLANCO, font=('Arial', 14, 'bold')).pack(side='left', padx=16)
        tk.Button(barra, text='📂  Abrir BC3', command=self._abrir_dialogo,
                  bg=COLOR_BLANCO, fg=COLOR_NARANJA, font=('Arial', 10, 'bold'),
                  relief='flat', padx=10, pady=3, cursor='hand2').pack(side='right', padx=12)

        self.lbl_archivo = tk.Label(self.root, text='Ningún archivo cargado',
                                    bg=COLOR_FONDO, fg='#666', font=('Arial', 9))
        self.lbl_archivo.pack(anchor='w', padx=16, pady=(6, 0))
        self.lbl_stats = tk.Label(self.root, text='', bg=COLOR_FONDO,
                                  fg=COLOR_NARANJA, font=('Arial', 9, 'bold'))
        self.lbl_stats.pack(anchor='w', padx=16)

        # Barra de búsqueda
        frm_busq = tk.Frame(self.root, bg=COLOR_FONDO)
        frm_busq.pack(fill='x', padx=16, pady=6)
        tk.Label(frm_busq, text='Buscar:', bg=COLOR_FONDO,
                 font=('Arial', 10)).pack(side='left')
        self.var_busq = tk.StringVar()
        self.var_busq.trace_add('write', lambda *_: self._buscar())
        tk.Entry(frm_busq, textvariable=self.var_busq, font=('Arial', 10),
                 width=38, relief='solid', bd=1).pack(side='left', padx=8)
        tk.Label(frm_busq, text='Ver:', bg=COLOR_FONDO,
                 font=('Arial', 10)).pack(side='left', padx=(12, 4))
        self.var_vista = tk.StringVar(value='conceptos')
        for txt, val in [('Conceptos', 'conceptos'), ('Textos', 'textos'),
                         ('Descomposiciones', 'descomp')]:
            tk.Radiobutton(frm_busq, text=txt, variable=self.var_vista,
                           value=val, bg=COLOR_FONDO,
                           command=self._actualizar_tabla).pack(side='left', padx=3)
        tk.Label(frm_busq, text='  Ctrl+clic o Shift+clic = selección múltiple',
                 bg=COLOR_FONDO, fg='#888', font=('Arial', 8)).pack(side='left')

        # Panel principal
        panel = tk.PanedWindow(self.root, orient='horizontal',
                               bg=COLOR_GRIS, sashwidth=5)
        panel.pack(fill='both', expand=True, padx=16, pady=(0, 0))

        # ── Tabla de conceptos ────────────────────────────────────────────────
        frm_tabla = tk.Frame(panel, bg=COLOR_FONDO)
        panel.add(frm_tabla, minsize=400)

        cols = ('codigo', 'unidad', 'resumen', 'precio')
        self.tabla = ttk.Treeview(frm_tabla, columns=cols, show='headings',
                                  selectmode='extended')
        self._cab_textos = {'codigo': 'Código', 'unidad': 'Ud.',
                            'resumen': 'Resumen', 'precio': 'Precio €'}
        for col, ancho, texto in [('codigo', 120, 'Código'), ('unidad', 60, 'Ud.'),
                                   ('resumen', 320, 'Resumen'), ('precio', 80, 'Precio €')]:
            self.tabla.heading(col, text=texto,
                               command=lambda c=col: self._ordenar(c))
            self.tabla.column(col, width=ancho, anchor='w')
        sy = ttk.Scrollbar(frm_tabla, orient='vertical', command=self.tabla.yview)
        sx = ttk.Scrollbar(frm_tabla, orient='horizontal', command=self.tabla.xview)
        self.tabla.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sy.pack(side='right', fill='y')
        sx.pack(side='bottom', fill='x')
        self.tabla.pack(fill='both', expand=True)
        self.tabla.tag_configure('par',   background='#F4F6F7')
        self.tabla.tag_configure('impar', background=COLOR_BLANCO)
        self.tabla.bind('<<TreeviewSelect>>', self._mostrar_detalle)
        self.tabla.bind('<Double-1>', self._anadir_al_carrito)

        frm_btn_add = tk.Frame(frm_tabla, bg=COLOR_FONDO, pady=4)
        frm_btn_add.pack(fill='x')
        tk.Button(frm_btn_add, text='➕  Añadir al carrito',
                  command=self._anadir_al_carrito,
                  bg=COLOR_VERDE, fg=COLOR_BLANCO, font=('Arial', 9, 'bold'),
                  relief='flat', padx=10, pady=4, cursor='hand2').pack(side='left', padx=4)
        tk.Label(frm_btn_add, text='(o doble clic en la fila)',
                 bg=COLOR_FONDO, fg='#888', font=('Arial', 8)).pack(side='left')

        # ── Detalle ───────────────────────────────────────────────────────────
        frm_det = tk.Frame(panel, bg=COLOR_BLANCO, padx=10, pady=10)
        panel.add(frm_det, minsize=230)
        tk.Label(frm_det, text='Detalle', bg=COLOR_BLANCO,
                 fg=COLOR_NARANJA, font=('Arial', 10, 'bold')).pack(anchor='w')
        self.txt_detalle = tk.Text(frm_det, wrap='word', font=('Consolas', 8),
                                   bg=COLOR_BLANCO, relief='flat',
                                   state='disabled', width=30)
        self.txt_detalle.pack(fill='both', expand=True, pady=(6, 0))

        # ── Carrito ───────────────────────────────────────────────────────────
        frm_carrito = tk.Frame(panel, bg=COLOR_FONDO, padx=6, pady=6)
        panel.add(frm_carrito, minsize=290)

        frm_cab = tk.Frame(frm_carrito, bg=COLOR_FONDO)
        frm_cab.pack(fill='x')
        tk.Label(frm_cab, text='🛒  Partidas seleccionadas',
                 bg=COLOR_FONDO, fg=COLOR_AZUL,
                 font=('Arial', 10, 'bold')).pack(side='left')
        self.lbl_n_carrito = tk.Label(frm_cab, text='(0)', bg=COLOR_FONDO,
                                      fg='#888', font=('Arial', 9))
        self.lbl_n_carrito.pack(side='left', padx=4)

        cols_c = ('cod', 'ud', 'resumen', 'precio', 'cantidad', 'importe')
        self.tbl_carrito = ttk.Treeview(frm_carrito, columns=cols_c,
                                        show='headings', selectmode='browse', height=18)
        for col, ancho, texto in [('cod', 90, 'Código'), ('ud', 40, 'Ud.'),
                                   ('resumen', 160, 'Descripción'), ('precio', 60, 'Precio'),
                                   ('cantidad', 55, 'Cant.'), ('importe', 65, 'Importe')]:
            self.tbl_carrito.heading(col, text=texto)
            self.tbl_carrito.column(col, width=ancho, anchor='w')
        sy_c = ttk.Scrollbar(frm_carrito, orient='vertical',
                              command=self.tbl_carrito.yview)
        self.tbl_carrito.configure(yscrollcommand=sy_c.set)
        sy_c.pack(side='right', fill='y')
        self.tbl_carrito.pack(fill='both', expand=True, pady=(4, 0))

        self.lbl_total = tk.Label(frm_carrito, text='Total: — €',
                                  bg=COLOR_FONDO, fg=COLOR_AZUL,
                                  font=('Arial', 10, 'bold'), anchor='e')
        self.lbl_total.pack(fill='x', pady=(2, 4))

        frm_btns = tk.Frame(frm_carrito, bg=COLOR_FONDO)
        frm_btns.pack(fill='x')

        tk.Button(frm_btns, text='📋 Copiar',
                  command=self._copiar_portapapeles,
                  bg=COLOR_AZUL, fg=COLOR_BLANCO, font=('Arial', 9, 'bold'),
                  relief='flat', padx=8, pady=4, cursor='hand2').pack(side='left', padx=2)

        if TIENE_OPENPYXL:
            tk.Button(frm_btns, text='📊 Excel',
                      command=self._exportar_excel,
                      bg=COLOR_NARANJA, fg=COLOR_BLANCO, font=('Arial', 9, 'bold'),
                      relief='flat', padx=8, pady=4, cursor='hand2').pack(side='left', padx=2)
        else:
            tk.Label(frm_btns, text='pip install openpyxl para xlsx',
                     bg=COLOR_FONDO, fg='#999', font=('Arial', 7)).pack(side='left')

        tk.Button(frm_btns, text='✏️ Cantidad',
                  command=self._editar_cantidad,
                  bg=COLOR_GRIS, fg=COLOR_TEXTO, font=('Arial', 9),
                  relief='flat', padx=6, pady=4, cursor='hand2').pack(side='right', padx=2)
        tk.Button(frm_btns, text='🗑 Quitar',
                  command=self._quitar_del_carrito,
                  bg='#AAAAAA', fg=COLOR_BLANCO, font=('Arial', 9),
                  relief='flat', padx=6, pady=4, cursor='hand2').pack(side='right', padx=2)

        # Barra de estado
        self.lbl_estado = tk.Label(self.root, text='', bg=COLOR_GRIS,
                                   fg=COLOR_TEXTO, font=('Arial', 8),
                                   anchor='w', padx=8)
        self.lbl_estado.pack(fill='x', side='bottom')

    # ── Carga ─────────────────────────────────────────────────────────────────

    def _abrir_dialogo(self):
        ruta = filedialog.askopenfilename(
            title='Abrir archivo BC3',
            filetypes=[('Archivos BC3', '*.bc3'), ('Todos', '*.*')])
        if ruta:
            self._cargar_archivo(ruta)

    def _cargar_archivo(self, ruta):
        try:
            self.lbl_estado.config(text=f'Cargando {ruta} …')
            self.root.update()
            self.pres = presupuesto(ruta)
            nombre = ruta.replace('\\', '/').split('/')[-1]
            self.lbl_archivo.config(text=f'📄  {nombre}')
            self.lbl_stats.config(
                text=(f'Conceptos: {len(self.pres.conceptos):,}   '
                      f'Descomposiciones: {len(self.pres.descomposiciones):,}   '
                      f'Textos: {len(self.pres.textos):,}   '
                      f'Importe total: {self.pres.importe_tot:,.2f} €'))
            self._actualizar_tabla()
            self.lbl_estado.config(text='Listo.')
        except Exception as e:
            messagebox.showerror('Error al cargar', str(e))
            self.lbl_estado.config(text='Error al cargar el archivo.')

    # ── Tabla ─────────────────────────────────────────────────────────────────

    def _actualizar_tabla(self):
        if self.pres:
            self._poblar_tabla(self.var_busq.get().strip().lower())

    def _poblar_tabla(self, filtro=''):
        for item in self.tabla.get_children():
            self.tabla.delete(item)
        vista = self.var_vista.get()
        if vista == 'conceptos':
            filas = self._filas_conceptos(filtro)
        elif vista == 'textos':
            filas = self._filas_textos(filtro)
        else:
            filas = self._filas_descomp(filtro)

        # Guardar cache y aplicar ordenación si hay una activa
        self._filas_cache = filas
        self._orden_col = None   # reset al cambiar filtro/vista
        self._resetear_cabeceras()
        self._volcar_filas(filas)
        self.lbl_estado.config(text=f'{len(filas):,} registros mostrados.')

    def _volcar_filas(self, filas):
        for item in self.tabla.get_children():
            self.tabla.delete(item)
        for i, fila in enumerate(filas):
            tag = 'par' if i % 2 == 0 else 'impar'
            self.tabla.insert('', 'end', values=fila, tags=(tag,))

    def _resetear_cabeceras(self):
        for col, texto in self._cab_textos.items():
            self.tabla.heading(col, text=texto)

    def _ordenar(self, col):
        """Ordena la tabla por la columna pulsada. Segundo clic invierte el orden."""
        idx = {'codigo': 0, 'unidad': 1, 'resumen': 2, 'precio': 3}[col]

        if self._orden_col == col:
            self._orden_asc = not self._orden_asc
        else:
            self._orden_col = col
            self._orden_asc = True

        # Ordenar: precio como número, resto como texto
        def clave(fila):
            v = fila[idx]
            if col == 'precio':
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return 0.0
            return str(v).lower()

        filas_ord = sorted(self._filas_cache, key=clave, reverse=not self._orden_asc)

        # Actualizar cabeceras con flecha
        self._resetear_cabeceras()
        flecha = ' ↑' if self._orden_asc else ' ↓'
        self.tabla.heading(col, text=self._cab_textos[col] + flecha)

        self._volcar_filas(filas_ord)
        self.lbl_estado.config(
            text=f'{len(filas_ord):,} registros — ordenados por {self._cab_textos[col]}{flecha}')

    def _filas_conceptos(self, filtro):
        filas = []
        for cod, datos in self.pres.conceptos.items():
            unidad  = datos[0] if len(datos) > 0 else ''
            resumen = datos[1] if len(datos) > 1 else ''
            precio  = datos[2] if len(datos) > 2 else ''
            if filtro and filtro not in cod.lower() and filtro not in resumen.lower():
                continue
            filas.append((cod, unidad, resumen, precio))
        return filas

    def _filas_textos(self, filtro):
        filas = []
        for cod, texto in self.pres.textos.items():
            corto = texto[:100].replace('\r', ' ').replace('\n', ' ')
            if filtro and filtro not in cod.lower() and filtro not in texto.lower():
                continue
            filas.append((cod, '', corto, ''))
        return filas

    def _filas_descomp(self, filtro):
        filas = []
        for cod, datos in self.pres.descomposiciones.items():
            contenido = datos[0] if datos else ''
            hijos = re.findall(r'(.*?)\\', contenido)[::3][:5]
            resumen = ', '.join(hijos)
            if filtro and filtro not in cod.lower() and filtro not in resumen.lower():
                continue
            filas.append((cod, '', resumen, ''))
        return filas

    def _buscar(self):
        self._orden_col = None
        self._actualizar_tabla()

    # ── Detalle ───────────────────────────────────────────────────────────────

    def _mostrar_detalle(self, event=None):
        sel = self.tabla.selection()
        if not sel:
            return
        valores = self.tabla.item(sel[-1])['values']
        codigo = str(valores[0])
        lineas = [f'CÓDIGO: {codigo}\n']
        if codigo in self.pres.conceptos:
            d = self.pres.conceptos[codigo]
            lineas += [f'Unidad:  {d[0] if len(d)>0 else ""}',
                       f'Resumen: {d[1] if len(d)>1 else ""}',
                       f'Precio:  {d[2] if len(d)>2 else ""} €',
                       f'Fecha:   {d[3] if len(d)>3 else ""}',
                       f'Tipo:    {d[4] if len(d)>4 else ""}', '']
        if codigo in self.pres.textos:
            lineas += ['── DESCRIPCIÓN ──', self.pres.textos[codigo], '']
        if codigo in self.pres.descomposiciones:
            lineas.append('── DESCOMPOSICIÓN ──')
            contenido = self.pres.descomposiciones[codigo][0] if self.pres.descomposiciones[codigo] else ''
            for h in re.findall(r'(.*?)\\(.*?)\\(.*?)\\', contenido):
                cod_h, _, cantidad = h
                ph = ''
                if cod_h in self.pres.conceptos and len(self.pres.conceptos[cod_h]) > 2:
                    ph = self.pres.conceptos[cod_h][2] + ' €'
                lineas.append(f'  {cod_h:20s}  x{cantidad:>8}  {ph}')
        texto = '\n'.join(str(l) for l in lineas)
        self.txt_detalle.config(state='normal')
        self.txt_detalle.delete('1.0', 'end')
        self.txt_detalle.insert('end', texto)
        self.txt_detalle.config(state='disabled')

    # ── Carrito ───────────────────────────────────────────────────────────────

    def _anadir_al_carrito(self, event=None):
        sel = self.tabla.selection()
        if not sel:
            return
        codigos_ya = {p['codigo'] for p in self.carrito}
        nuevos = 0
        for item in sel:
            valores = self.tabla.item(item)['values']
            cod = str(valores[0])
            if cod in codigos_ya:
                continue
            datos = self.pres.conceptos.get(cod, [])
            unidad      = datos[0] if len(datos) > 0 else ''
            resumen     = datos[1] if len(datos) > 1 else ''
            precio_str  = datos[2] if len(datos) > 2 else '0'
            try:
                precio = float(precio_str)
            except ValueError:
                precio = 0.0
            descripcion = self.pres.textos.get(cod, resumen)
            self.carrito.append({'codigo': cod, 'unidad': unidad, 'resumen': resumen,
                                 'descripcion': descripcion, 'precio': precio, 'cantidad': 1.0})
            nuevos += 1
        if nuevos:
            self._refrescar_carrito()
            self.lbl_estado.config(text=f'{nuevos} partida(s) añadida(s) al carrito.')
        else:
            self.lbl_estado.config(text='Las partidas ya estaban en el carrito.')

    def _refrescar_carrito(self):
        for item in self.tbl_carrito.get_children():
            self.tbl_carrito.delete(item)
        total = 0.0
        for p in self.carrito:
            importe = p['precio'] * p['cantidad']
            total += importe
            self.tbl_carrito.insert('', 'end', values=(
                p['codigo'], p['unidad'], p['resumen'][:30],
                f"{p['precio']:.2f}", f"{p['cantidad']:.2f}", f"{importe:.2f}"))
        self.lbl_n_carrito.config(text=f'({len(self.carrito)})')
        self.lbl_total.config(text=f'Total: {total:,.2f} €' if self.carrito else 'Total: — €')

    def _quitar_del_carrito(self):
        sel = self.tbl_carrito.selection()
        if not sel:
            return
        idx = self.tbl_carrito.index(sel[0])
        del self.carrito[idx]
        self._refrescar_carrito()

    def _editar_cantidad(self):
        sel = self.tbl_carrito.selection()
        if not sel:
            messagebox.showinfo('Editar cantidad', 'Selecciona primero una partida del carrito.')
            return
        idx = self.tbl_carrito.index(sel[0])
        partida = self.carrito[idx]

        v = tk.Toplevel(self.root)
        v.title('Editar cantidad')
        v.geometry('300x130')
        v.resizable(False, False)
        v.grab_set()
        tk.Label(v, text=f'{partida["codigo"]} — {partida["resumen"][:40]}',
                 wraplength=280, justify='left').pack(padx=12, pady=(12, 4))
        var_cant = tk.StringVar(value=str(partida['cantidad']))
        entry = tk.Entry(v, textvariable=var_cant, font=('Arial', 11), width=12)
        entry.pack(pady=4)
        entry.select_range(0, 'end')
        entry.focus()

        def aplicar():
            try:
                partida['cantidad'] = float(var_cant.get().replace(',', '.'))
                self._refrescar_carrito()
                v.destroy()
            except ValueError:
                messagebox.showerror('Error', 'Introduce un número válido.')

        tk.Button(v, text='Aceptar', command=aplicar,
                  bg=COLOR_NARANJA, fg=COLOR_BLANCO, font=('Arial', 10, 'bold'),
                  relief='flat', padx=10, pady=4).pack(pady=6)
        v.bind('<Return>', lambda e: aplicar())

    # ── Exportar ──────────────────────────────────────────────────────────────

    def _copiar_portapapeles(self):
        if not self.carrito:
            messagebox.showinfo('Carrito vacío', 'Añade partidas al carrito primero.')
            return
        lineas = ['Código\tUnidad\tDescripción\tPrecio €\tCantidad\tImporte €']
        total = 0.0
        for p in self.carrito:
            importe = p['precio'] * p['cantidad']
            total += importe
            lineas.append(f"{p['codigo']}\t{p['unidad']}\t{p['resumen']}\t"
                          f"{p['precio']:.2f}\t{p['cantidad']:.2f}\t{importe:.2f}")
        lineas.append(f"\t\t\t\tTOTAL\t{total:.2f}")
        self.root.clipboard_clear()
        self.root.clipboard_append('\n'.join(lineas))
        self.lbl_estado.config(
            text=f'✅  {len(self.carrito)} partidas copiadas. Pega con Ctrl+V en Excel.')

    def _exportar_excel(self):
        if not self.carrito:
            messagebox.showinfo('Carrito vacío', 'Añade partidas al carrito primero.')
            return
        ruta = filedialog.asksaveasfilename(
            title='Guardar como Excel',
            defaultextension='.xlsx',
            filetypes=[('Excel', '*.xlsx')])
        if not ruta:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Partidas BC3'

            naranja = 'E87722'
            azul    = '1A5276'
            fill_cab = PatternFill('solid', start_color=naranja)
            fill_tot = PatternFill('solid', start_color=azul)
            fill_alt = PatternFill('solid', start_color='F4F6F7')
            font_cab = Font(bold=True, color='FFFFFF', size=10)
            font_tot = Font(bold=True, color='FFFFFF', size=10)
            aln_c    = Alignment(horizontal='center')
            aln_r    = Alignment(horizontal='right')

            # Cabecera empresa
            ws.merge_cells('A1:F1')
            ws['A1'] = 'Excavaciones Arturo S.L. — Partidas de presupuesto'
            ws['A1'].font = Font(bold=True, size=12, color=naranja)
            ws.merge_cells('A2:F2')
            ws['A2'] = self.lbl_archivo.cget('text').replace('📄  ', '')
            ws['A2'].font = Font(italic=True, size=9, color='888888')

            # Cabeceras tabla
            cabeceras = ['Código', 'Unidad', 'Descripción completa',
                         'Precio ud. (€)', 'Cantidad', 'Importe (€)']
            for col, cab in enumerate(cabeceras, 1):
                c = ws.cell(row=4, column=col, value=cab)
                c.font = font_cab
                c.fill = fill_cab
                c.alignment = aln_c

            # Datos
            fi = 5
            for i, p in enumerate(self.carrito):
                fila = fi + i
                ws.cell(fila, 1, p['codigo'])
                ws.cell(fila, 2, p['unidad']).alignment = aln_c
                ws.cell(fila, 3, p['descripcion'] or p['resumen'])
                c4 = ws.cell(fila, 4, p['precio'])
                c5 = ws.cell(fila, 5, p['cantidad'])
                c6 = ws.cell(fila, 6)
                c4.number_format = c5.number_format = c6.number_format = '#,##0.00'
                c4.alignment = c5.alignment = c6.alignment = aln_r
                c6.value = f'=D{fila}*E{fila}'
                if i % 2 == 0:
                    for col in range(1, 7):
                        ws.cell(fila, col).fill = fill_alt

            # Fila total
            ft = fi + len(self.carrito)
            ws.merge_cells(f'A{ft}:E{ft}')
            ws[f'A{ft}'] = 'TOTAL PRESUPUESTO'
            ws[f'A{ft}'].font = font_tot
            ws[f'A{ft}'].fill = fill_tot
            ws[f'A{ft}'].alignment = aln_r
            c_tot = ws.cell(ft, 6, f'=SUM(F{fi}:F{ft-1})')
            c_tot.number_format = '#,##0.00'
            c_tot.font = font_tot
            c_tot.fill = fill_tot
            c_tot.alignment = aln_r

            # Anchos
            for col, ancho in [(1,16),(2,8),(3,55),(4,14),(5,12),(6,14)]:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho
            ws.freeze_panes = 'A5'

            wb.save(ruta)
            self.lbl_estado.config(text=f'✅  Excel guardado: {ruta}')
            messagebox.showinfo('Exportado', f'Archivo guardado:\n{ruta}')
        except Exception as e:
            messagebox.showerror('Error al exportar', str(e))


if __name__ == '__main__':
    root = tk.Tk()
    app = ExploradorBC3(root)
    root.mainloop()